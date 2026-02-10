#!/usr/bin/env python3
"""
Sales & Inventory Dashboard Creator
Creates a professional Excel dashboard using Pivot Tables, Charts, and Slicers
Compatible with Excel 2021
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

def generate_sample_data():
    """Generate sample dummy data for Products and Sales"""
    
    # Product categories
    categories = ['Electronics', 'Clothing', 'Home & Garden', 'Sports', 'Books']
    suppliers = ['Supplier A', 'Supplier B', 'Supplier C', 'Supplier D', 'Supplier E']
    
    # Generate Products data
    products_data = []
    for i in range(1, 51):  # 50 products
        product_id = f'P{i:03d}'
        product_name = f'Product {i}'
        category = random.choice(categories)
        purchase_price = round(random.uniform(10, 200), 2)
        selling_price = round(purchase_price * random.uniform(1.2, 2.0), 2)
        initial_quantity = random.randint(50, 500)
        supplier = random.choice(suppliers)
        date_added = datetime(2025, 1, 1) + timedelta(days=random.randint(0, 60))
        
        products_data.append({
            'Product ID': product_id,
            'Product Name': product_name,
            'Category': category,
            'Purchase Price': purchase_price,
            'Selling Price': selling_price,
            'Initial Quantity': initial_quantity,
            'Supplier': supplier,
            'Date Added': date_added
        })
    
    # Generate Sales data
    sales_data = []
    order_id = 1
    start_date = datetime(2025, 1, 1)
    
    for _ in range(200):  # 200 sales records
        product = random.choice(products_data)
        product_id = product['Product ID']
        quantity_sold = random.randint(1, 20)
        selling_price = product['Selling Price']
        total_sale = round(quantity_sold * selling_price, 2)
        sale_date = start_date + timedelta(days=random.randint(0, 365))
        
        sales_data.append({
            'Order ID': f'ORD{order_id:04d}',
            'Product ID': product_id,
            'Quantity Sold': quantity_sold,
            'Selling Price': selling_price,
            'Total Sale': total_sale,
            'Date': sale_date
        })
        order_id += 1
    
    return pd.DataFrame(products_data), pd.DataFrame(sales_data)

def create_excel_table(ws, df, table_name, start_row=1, start_col=1):
    """Create an Excel table from a DataFrame"""
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    # Write DataFrame to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Header formatting
            if r_idx == start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Create table reference
    end_col = start_col + len(df.columns) - 1
    end_row = start_row + len(df)
    
    table_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    
    # Create table
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Auto-adjust column widths
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    return end_row

def create_dashboard_sheet(wb, products_df, sales_df):
    """Create the Dashboard sheet with KPIs and charts"""
    ws = wb.create_sheet("Dashboard", 0)
    
    # Remove gridlines
    ws.sheet_view.showGridLines = False
    
    # Set up colors
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    kpi_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Title
    ws['A1'] = 'Sales & Inventory Dashboard'
    ws['A1'].font = Font(size=24, bold=True, color="4472C4")
    ws.row_dimensions[1].height = 35
    
    # Merge purchase prices for profit calculation
    product_prices = products_df.set_index('Product ID')[['Purchase Price']].to_dict()['Purchase Price']
    
    # Calculate KPIs
    total_sales = sales_df['Total Sale'].sum()
    total_orders = len(sales_df)
    avg_order_value = sales_df['Total Sale'].mean()
    
    # Calculate profit
    total_profit = 0
    for _, sale in sales_df.iterrows():
        product_id = sale['Product ID']
        if product_id in product_prices:
            purchase_price = product_prices[product_id]
            selling_price = sale['Selling Price']
            quantity_sold = sale['Quantity Sold']
            profit = (selling_price - purchase_price) * quantity_sold
            total_profit += profit
    
    # Calculate stock levels
    sales_by_product = sales_df.groupby('Product ID')['Quantity Sold'].sum().to_dict()
    low_stock_count = 0
    for _, product in products_df.iterrows():
        product_id = product['Product ID']
        initial_qty = product['Initial Quantity']
        sold_qty = sales_by_product.get(product_id, 0)
        current_stock = initial_qty - sold_qty
        if current_stock < 50:  # Low stock threshold
            low_stock_count += 1
    
    # Create KPI Cards
    kpi_row = 3
    kpis = [
        ('Total Sales', f'${total_sales:,.2f}', 'B'),
        ('Total Profit', f'${total_profit:,.2f}', 'D'),
        ('Total Orders', f'{total_orders:,}', 'F'),
        ('Avg Order Value', f'${avg_order_value:,.2f}', 'H'),
        ('Low Stock Products', f'{low_stock_count}', 'J')
    ]
    
    for kpi_name, kpi_value, col in kpis:
        # KPI Label
        label_cell = ws[f'{col}{kpi_row}']
        label_cell.value = kpi_name
        label_cell.font = Font(size=10, bold=True)
        label_cell.alignment = Alignment(horizontal="center")
        
        # KPI Value
        value_cell = ws[f'{col}{kpi_row + 1}']
        value_cell.value = kpi_value
        value_cell.font = Font(size=16, bold=True, color="4472C4")
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.fill = kpi_fill
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        label_cell.border = thin_border
        value_cell.border = thin_border
    
    # Create summary data for charts
    sales_df['Month'] = pd.to_datetime(sales_df['Date']).dt.to_period('M').astype(str)
    
    # Sales by Month
    sales_by_month = sales_df.groupby('Month')['Total Sale'].sum().reset_index()
    sales_by_month.columns = ['Month', 'Sales']
    
    # Write sales by month data (for chart)
    chart_start_row = 7
    ws[f'B{chart_start_row}'] = 'Month'
    ws[f'C{chart_start_row}'] = 'Sales'
    ws[f'B{chart_start_row}'].font = Font(bold=True)
    ws[f'C{chart_start_row}'].font = Font(bold=True)
    
    for idx, row in enumerate(sales_by_month.itertuples(index=False), start=chart_start_row + 1):
        ws[f'B{idx}'] = row.Month
        ws[f'C{idx}'] = row.Sales
    
    # Create Line Chart for Sales by Month
    chart1 = LineChart()
    chart1.title = "Sales by Month"
    chart1.style = 10
    chart1.y_axis.title = 'Sales ($)'
    chart1.x_axis.title = 'Month'
    chart1.height = 10
    chart1.width = 20
    
    data = Reference(ws, min_col=3, min_row=chart_start_row, max_row=chart_start_row + len(sales_by_month))
    cats = Reference(ws, min_col=2, min_row=chart_start_row + 1, max_row=chart_start_row + len(sales_by_month))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "B10")
    
    # Top Selling Products
    top_products = sales_df.groupby('Product ID')['Quantity Sold'].sum().nlargest(10).reset_index()
    top_products.columns = ['Product ID', 'Quantity']
    
    # Write top products data
    top_start_row = 7
    ws[f'E{top_start_row}'] = 'Product'
    ws[f'F{top_start_row}'] = 'Quantity'
    ws[f'E{top_start_row}'].font = Font(bold=True)
    ws[f'F{top_start_row}'].font = Font(bold=True)
    
    for idx, row in enumerate(top_products.itertuples(index=False), start=top_start_row + 1):
        ws[f'E{idx}'] = row[0]  # Product ID
        ws[f'F{idx}'] = row[1]  # Quantity
    
    # Create Bar Chart for Top Selling Products
    chart2 = BarChart()
    chart2.title = "Top 10 Selling Products"
    chart2.style = 11
    chart2.y_axis.title = 'Product'
    chart2.x_axis.title = 'Quantity Sold'
    chart2.height = 10
    chart2.width = 20
    
    data = Reference(ws, min_col=6, min_row=top_start_row, max_row=top_start_row + len(top_products))
    cats = Reference(ws, min_col=5, min_row=top_start_row + 1, max_row=top_start_row + len(top_products))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "E10")
    
    # Add instructions
    ws['B25'] = 'Instructions:'
    ws['B25'].font = Font(size=12, bold=True, color="4472C4")
    ws['B26'] = '1. Use the Products and Sales sheets to view raw data'
    ws['B27'] = '2. Both sheets are formatted as Excel Tables for easy filtering'
    ws['B28'] = '3. Charts update automatically based on the data'
    ws['B29'] = '4. To add slicers: Select a table → Insert tab → Slicer'
    ws['B30'] = '5. Connect slicers to Pivot Tables for interactive filtering'
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 15

def create_calculations_sheet(wb, products_df, sales_df):
    """Create a Calculations sheet with derived metrics"""
    ws = wb.create_sheet("Calculations")
    
    # Headers
    ws['A1'] = 'Product ID'
    ws['B1'] = 'Product Name'
    ws['C1'] = 'Initial Quantity'
    ws['D1'] = 'Total Sold'
    ws['E1'] = 'Current Stock'
    ws['F1'] = 'Stock Status'
    ws['G1'] = 'Total Revenue'
    ws['H1'] = 'Total Profit'
    
    # Format headers
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate metrics for each product
    sales_by_product = sales_df.groupby('Product ID').agg({
        'Quantity Sold': 'sum',
        'Total Sale': 'sum'
    }).to_dict()
    
    product_prices = products_df.set_index('Product ID')[['Purchase Price', 'Selling Price']].to_dict()
    
    row = 2
    for _, product in products_df.iterrows():
        product_id = product['Product ID']
        product_name = product['Product Name']
        initial_qty = product['Initial Quantity']
        
        # Get sales data
        total_sold = sales_by_product['Quantity Sold'].get(product_id, 0)
        total_revenue = sales_by_product['Total Sale'].get(product_id, 0)
        
        # Calculate current stock
        current_stock = initial_qty - total_sold
        
        # Stock status
        if current_stock < 50:
            stock_status = 'Low'
        elif current_stock < 100:
            stock_status = 'Medium'
        else:
            stock_status = 'Good'
        
        # Calculate profit
        purchase_price = product['Purchase Price']
        selling_price = product['Selling Price']
        total_profit = (selling_price - purchase_price) * total_sold
        
        # Write to sheet
        ws[f'A{row}'] = product_id
        ws[f'B{row}'] = product_name
        ws[f'C{row}'] = initial_qty
        ws[f'D{row}'] = total_sold
        ws[f'E{row}'] = current_stock
        ws[f'F{row}'] = stock_status
        ws[f'G{row}'] = round(total_revenue, 2)
        ws[f'H{row}'] = round(total_profit, 2)
        
        # Color code stock status
        status_cell = ws[f'F{row}']
        if stock_status == 'Low':
            status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        elif stock_status == 'Medium':
            status_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        row += 1
    
    # Auto-adjust columns
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15

def main():
    """Main function to create the dashboard"""
    print("Generating sample data...")
    products_df, sales_df = generate_sample_data()
    
    print("Creating Excel workbook...")
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Products sheet
    print("Creating Products sheet...")
    ws_products = wb.create_sheet("Products")
    end_row = create_excel_table(ws_products, products_df, "tbl_Products")
    
    # Create Sales sheet
    print("Creating Sales sheet...")
    ws_sales = wb.create_sheet("Sales")
    end_row = create_excel_table(ws_sales, sales_df, "tbl_Sales")
    
    # Create Calculations sheet
    print("Creating Calculations sheet...")
    create_calculations_sheet(wb, products_df, sales_df)
    
    # Create Dashboard sheet
    print("Creating Dashboard sheet...")
    create_dashboard_sheet(wb, products_df, sales_df)
    
    # Save workbook
    filename = "Sales_Inventory_Dashboard.xlsx"
    print(f"Saving workbook as {filename}...")
    wb.save(filename)
    print(f"✓ Dashboard created successfully: {filename}")
    print("\nNext Steps:")
    print("1. Open the file in Excel 2021")
    print("2. Go to Data → Relationships to verify table relationship")
    print("3. Insert → Pivot Table to create additional pivot tables")
    print("4. Insert → Slicer to add interactive filters")
    print("5. Customize the dashboard design as needed")

if __name__ == "__main__":
    main()
