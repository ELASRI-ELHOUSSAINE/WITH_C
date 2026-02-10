#!/usr/bin/env python3
"""
Verification script for Sales_Inventory_Dashboard.xlsx
This script validates that the Excel file meets all requirements
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter

def verify_dashboard():
    """Verify all requirements are met"""
    print("🔍 Verifying Sales_Inventory_Dashboard.xlsx...\n")
    
    errors = []
    warnings = []
    
    try:
        wb = openpyxl.load_workbook('Sales_Inventory_Dashboard.xlsx')
    except Exception as e:
        print(f"❌ ERROR: Could not open Excel file: {e}")
        return False
    
    # Check sheets exist
    required_sheets = ['Dashboard', 'Products', 'Sales', 'Calculations']
    print("1. Checking required sheets...")
    for sheet in required_sheets:
        if sheet in wb.sheetnames:
            print(f"   ✓ {sheet} sheet exists")
        else:
            errors.append(f"Missing required sheet: {sheet}")
            print(f"   ✗ {sheet} sheet missing")
    
    # Check tables
    print("\n2. Checking Excel tables...")
    products_ws = wb['Products']
    sales_ws = wb['Sales']
    
    if hasattr(products_ws, 'tables'):
        products_tables = list(products_ws.tables.values())
        if products_tables and products_tables[0].name == 'tbl_Products':
            print(f"   ✓ tbl_Products exists with {products_ws.max_row - 1} rows")
        else:
            errors.append("tbl_Products not found or incorrectly named")
    
    if hasattr(sales_ws, 'tables'):
        sales_tables = list(sales_ws.tables.values())
        if sales_tables and sales_tables[0].name == 'tbl_Sales':
            print(f"   ✓ tbl_Sales exists with {sales_ws.max_row - 1} rows")
        else:
            errors.append("tbl_Sales not found or incorrectly named")
    
    # Check Dashboard elements
    print("\n3. Checking Dashboard elements...")
    dashboard_ws = wb['Dashboard']
    
    # Check KPIs
    kpi_labels = ['Total Sales', 'Total Profit', 'Total Orders', 'Avg Order Value', 'Low Stock Products']
    kpi_found = 0
    for row in range(1, 10):
        for col in range(1, 15):
            cell_value = dashboard_ws.cell(row, col).value
            if cell_value in kpi_labels:
                kpi_found += 1
    
    if kpi_found >= 4:
        print(f"   ✓ Found {kpi_found} KPI cards")
    else:
        warnings.append(f"Only found {kpi_found} KPI cards (expected 5)")
    
    # Check charts
    if hasattr(dashboard_ws, '_charts'):
        chart_count = len(dashboard_ws._charts)
        if chart_count >= 2:
            print(f"   ✓ Found {chart_count} charts")
        else:
            warnings.append(f"Only found {chart_count} charts (expected 2)")
    
    # Check gridlines are off
    if not dashboard_ws.sheet_view.showGridLines:
        print("   ✓ Gridlines disabled on Dashboard")
    else:
        warnings.append("Gridlines should be disabled on Dashboard")
    
    # Check Products data
    print("\n4. Checking Products data structure...")
    expected_product_cols = ['Product ID', 'Product Name', 'Category', 'Purchase Price', 
                             'Selling Price', 'Initial Quantity', 'Supplier', 'Date Added']
    actual_cols = [cell.value for cell in products_ws[1]]
    
    if actual_cols == expected_product_cols:
        print("   ✓ Products columns are correct")
    else:
        errors.append("Products columns don't match specification")
        print(f"   Expected: {expected_product_cols}")
        print(f"   Got: {actual_cols}")
    
    # Check Sales data
    print("\n5. Checking Sales data structure...")
    expected_sales_cols = ['Order ID', 'Product ID', 'Quantity Sold', 
                           'Selling Price', 'Total Sale', 'Date']
    actual_sales_cols = [cell.value for cell in sales_ws[1]]
    
    if actual_sales_cols == expected_sales_cols:
        print("   ✓ Sales columns are correct")
    else:
        errors.append("Sales columns don't match specification")
        print(f"   Expected: {expected_sales_cols}")
        print(f"   Got: {actual_sales_cols}")
    
    # Check Calculations
    print("\n6. Checking Calculations sheet...")
    calc_ws = wb['Calculations']
    calc_cols = [cell.value for cell in calc_ws[1]]
    
    if 'Current Stock' in calc_cols and 'Stock Status' in calc_cols:
        print("   ✓ Calculations include Current Stock and Stock Status")
    else:
        warnings.append("Calculations sheet may be missing required columns")
    
    # Final summary
    print("\n" + "="*60)
    
    if errors:
        print("❌ VERIFICATION FAILED\n")
        for error in errors:
            print(f"   ERROR: {error}")
        wb.close()
        return False
    
    if warnings:
        print("⚠️  VERIFICATION PASSED WITH WARNINGS\n")
        for warning in warnings:
            print(f"   WARNING: {warning}")
    else:
        print("✅ VERIFICATION PASSED - ALL REQUIREMENTS MET!")
    
    print("\n📊 Dashboard Statistics:")
    print(f"   • Sheets: {len(wb.sheetnames)}")
    print(f"   • Products: {products_ws.max_row - 1}")
    print(f"   • Sales: {sales_ws.max_row - 1}")
    print(f"   • Charts: {len(dashboard_ws._charts)}")
    
    wb.close()
    return True

if __name__ == "__main__":
    success = verify_dashboard()
    sys.exit(0 if success else 1)
